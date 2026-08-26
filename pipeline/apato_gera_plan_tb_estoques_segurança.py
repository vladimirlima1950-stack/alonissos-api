import os
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers

def run(pasta_cliente):

    pasta_saida = os.path.join(pasta_cliente, "saida")
    os.makedirs(pasta_saida, exist_ok=True)

    caminho_duckdb = os.path.join(pasta_cliente, "processamento", "previsao.duckdb")

    con = duckdb.connect(caminho_duckdb, read_only=False)

    # ============================================================
    # CONSULTA AO BANCO
    # ============================================================
    df = con.execute("""
        SELECT class_abc_pedidos, class_abc_valores, sku,
               es_mes_corrente, es_corr_mais1, es_corr_mais2, es_corr_mais3,
               es_corr_mais4, es_corr_mais5, es_corr_mais6, es_corr_mais7,
               es_corr_mais8, es_corr_mais9, es_corr_mais10, es_corr_mais11
        FROM tb_estoq_segur_12meses_res_fim
    """).df()

    # ============================================================
    # TÍTULOS E CAMPOS (mantidos exatamente como no PHP)
    # ============================================================
    titulos = [
        'frequência de pedidos', 'importância no faturamento', 'sku número',
        'ES mês corrente', 'ES mês corrente+1', 'ES mês corrente+2', 'ES mês corrente+3',
        'ES mês corrente+4', 'ES mês corrente+5', 'ES mês corrente+6', 'ES mês corrente+7',
        'ES mês corrente+8', 'ES mês corrente+9', 'ES mês corrente+10', 'ES mês corrente+11'
    ]

    campos = [
        'class_abc_pedidos', 'class_abc_valores', 'sku',
        'es_mes_corrente', 'es_corr_mais1', 'es_corr_mais2', 'es_corr_mais3',
        'es_corr_mais4', 'es_corr_mais5', 'es_corr_mais6', 'es_corr_mais7',
        'es_corr_mais8', 'es_corr_mais9', 'es_corr_mais10', 'es_corr_mais11'
    ]

    # ============================================================
    # CRIAÇÃO DA PLANILHA
    # ============================================================
    wb = Workbook()
    ws = wb.active

    # Cabeçalhos
    for idx, titulo in enumerate(titulos, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = titulo
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 15

    # ============================================================
    # PREENCHENDO DADOS
    # ============================================================
    linha = 2

    for _, row in df.iterrows():
        for col_idx, campo in enumerate(campos, start=1):

            valor = row[campo]
            cell = ws.cell(row=linha, column=col_idx)

            # Alinhamento igual ao PHP
            if campo == "sku":
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

            # Valor
            if isinstance(valor, (int, float)):
                cell.value = float(valor)
            else:
                cell.value = valor

            # Formatação numérica igual ao PHP
            if campo.startswith("es_") and isinstance(valor, (int, float)):
                cell.number_format = numbers.FORMAT_NUMBER

        linha += 1

    # ============================================================
    # SALVA ARQUIVO
    # ============================================================
    caminho_arquivo = os.path.join(pasta_saida, "tabela_estoques_segurança.xlsx")
    wb.save(caminho_arquivo)

    print(f"Planilha gerada com sucesso em: {caminho_arquivo}")

    return caminho_arquivo
