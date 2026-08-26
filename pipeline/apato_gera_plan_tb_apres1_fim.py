import os
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers

def run(pasta_cliente):

    pasta_saida = os.path.join(pasta_cliente, "saida")
    os.makedirs(pasta_saida, exist_ok=True)

    caminho_duckdb = os.path.join(pasta_cliente, "processamento", "previsao.duckdb")

    con = duckdb.connect(caminho_duckdb, read_only=False)

    # ============================================================
    # CONSULTA AO BANCO
    # ============================================================
    df = con.execute("""
        SELECT sku, ativo_inativo, class_pedidos, class_custo,
               qtde_pedidos, porc_pedidos, custo_12meses,
               porc_custo_12meses, custo_estoque, porc_custo_estoque,
               giro_estoque_sku, cobertura_semanas_sku
        FROM tb_apres1_fim
    """).df()

    # ============================================================
    # CRIA PLANILHA
    # ============================================================
    wb = Workbook()
    ws = wb.active

    headers = [
        'sku número', 'ativo_inativo', 'frequência de pedidos', 'importância no faturamento',
        'quantidade de pedidos em 12 meses', 'porcentagem de pedidos em 12 meses',
        'custo do produto em 12 meses', 'porcentagem do custo em 12 meses',
        'custo do estoque', 'porcentagem do custo do estoque',
        'giro do estoque', 'cobertura em semanas'
    ]

    ws.append(headers)

    # Estilo do cabeçalho
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    # ============================================================
    # INSERE DADOS
    # ============================================================
    for _, row in df.iterrows():
        ws.append([
            row["sku"],
            row["ativo_inativo"],
            row["class_pedidos"],
            row["class_custo"],
            int(row["qtde_pedidos"]),
            float(row["porc_pedidos"]),
            float(row["custo_12meses"]),
            float(row["porc_custo_12meses"]),
            float(row["custo_estoque"]),
            float(row["porc_custo_estoque"]),
            float(row["giro_estoque_sku"]),
            float(row["cobertura_semanas_sku"])
        ])

    # ============================================================
    # FORMATOS NUMÉRICOS
    # ============================================================
    totalLinhas = len(df) + 1

    for row in range(2, totalLinhas + 1):
        ws[f"E{row}"].number_format = numbers.FORMAT_NUMBER
        ws[f"F{row}"].number_format = numbers.FORMAT_PERCENTAGE_00
        ws[f"G{row}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        ws[f"H{row}"].number_format = numbers.FORMAT_PERCENTAGE_00
        ws[f"I{row}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        ws[f"J{row}"].number_format = numbers.FORMAT_PERCENTAGE_00
        ws[f"K{row}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        ws[f"L{row}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # ============================================================
    # SALVA ARQUIVO
    # ============================================================
    caminho_arquivo = os.path.join(pasta_saida, "tabela_apres1.xlsx")
    wb.save(caminho_arquivo)

    print(f"Planilha gerada com sucesso em: {caminho_arquivo}")

    return caminho_arquivo
