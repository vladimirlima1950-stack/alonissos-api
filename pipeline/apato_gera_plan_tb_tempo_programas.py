import os
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers

def run(pasta_cliente):

    pasta_saida = os.path.join(pasta_cliente, "saida")
    os.makedirs(pasta_saida, exist_ok=True)

    caminho_duckdb = os.path.join(pasta_cliente, "processamento", "previsao.duckdb")
    caminho_arquivo = os.path.join(pasta_saida, "tabela_tempo_programa.xlsx")

    con = duckdb.connect(caminho_duckdb, read_only=False)

    # ============================================================
    # CONSULTA AO BANCO
    # ============================================================
    df = con.execute("""
        SELECT nome_programa, inicio, fim, evento
        FROM tb_tempo_programa
    """).df()

    # ============================================================
    # CRIAÇÃO DA PLANILHA
    # ============================================================
    wb = Workbook()
    ws = wb.active

    cabecalhos = ['nome_programa', 'inicio', 'fim', 'evento']
    ws.append(cabecalhos)

    # Estilo dos cabeçalhos
    for col_idx in range(1, 5):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 30

    # ============================================================
    # INSERIR DADOS
    # ============================================================
    for _, row in df.iterrows():
        ws.append([
            row["nome_programa"],
            row["inicio"],
            row["fim"],
            row["evento"]
        ])

    totalLinhas = len(df) + 1

    # Formatação de datas (igual ao PHP)
    if totalLinhas > 1:
        ws[f"B2:B{totalLinhas}"].number_format = numbers.FORMAT_DATE_DATETIME
        ws[f"C2:C{totalLinhas}"].number_format = numbers.FORMAT_DATE_DATETIME

    # ============================================================
    # SALVA ARQUIVO
    # ============================================================
    wb.save(caminho_arquivo)

    print(f"Planilha de tempos de execução gerada com sucesso em: {caminho_arquivo}")

    return caminho_arquivo
