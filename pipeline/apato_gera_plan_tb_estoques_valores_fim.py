import os
import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)

# ============================================================
# Função para localizar colunas ignorando maiúsculas/minúsculas
# ============================================================
def encontrar_coluna(con, tabela, palavra):
    palavra = palavra.lower()
    cols = con.execute(f"DESCRIBE {tabela}").fetchdf()["column_name"].tolist()

    for c in cols:
        if palavra in c.lower():
            return c

    raise Exception(f"Nenhuma coluna contendo '{palavra}' encontrada na tabela {tabela}.")


# ============================================================
# Função principal
# ============================================================
def run(pasta_cliente):

    pasta_saida = os.path.join(pasta_cliente, "saida")
    os.makedirs(pasta_saida, exist_ok=True)

    pasta_processamento = os.path.join(pasta_cliente, "processamento")
    caminho_banco = os.path.join(pasta_processamento, "previsao.duckdb")

    print("Conectando ao DuckDB:", caminho_banco)
    con = duckdb.connect(caminho_banco, read_only=False)

    tabela = "tb_estoques_valores_fim"

    # ============================================================
    # Localiza automaticamente as colunas (case-insensitive)
    # ============================================================
    col_class_ped = encontrar_coluna(con, tabela, "class_abc_ped")
    col_class_val = encontrar_coluna(con, tabela, "class_abc_val")
    col_sku       = encontrar_coluna(con, tabela, "sku")
    col_qtde      = encontrar_coluna(con, tabela, "qtde")
    col_custo     = encontrar_coluna(con, tabela, "custo_unit")
    col_total     = encontrar_coluna(con, tabela, "total_valor_sku")
    col_porc      = encontrar_coluna(con, tabela, "porc_valor_estoque")

    # ============================================================
    # Consulta ao banco usando os nomes reais das colunas
    # ============================================================
    df = con.execute(f"""
        SELECT 
            {col_class_ped},
            {col_class_val},
            {col_sku},
            {col_qtde},
            {col_custo},
            {col_total},
            {col_porc}
        FROM {tabela}
    """).df()

    # ============================================================
    # Títulos e campos (mantidos como no PHP)
    # ============================================================
    titulos = [
        'frequência de pedidos', 'importância no faturamento', 'sku número',
        'quantidade no estoque', 'custo unitário', 'valor total em estoque',
        'porcentagem do valor do estoque'
    ]

    campos = [
        col_class_ped, col_class_val, col_sku,
        col_qtde, col_custo, col_total, col_porc
    ]

    # ============================================================
    # Criação da planilha
    # ============================================================
    wb = Workbook()
    ws = wb.active

    # Cabeçalhos
    for idx, titulo in enumerate(titulos, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = titulo
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 18

    # ============================================================
    # Preenchendo dados
    # ============================================================
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2
        for col_idx, campo in enumerate(campos, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            valor = row[campo]

            # Alinhamento
            if campo == col_sku:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

            # Valor
            cell.value = float(valor) if isinstance(valor, (int, float)) else valor

    # ============================================================
    # Salva a planilha
    # ============================================================
    caminho_planilha = os.path.join(pasta_saida, "tabela_estoques_valores.xlsx")
    wb.save(caminho_planilha)

    print(f"Planilha gerada com sucesso em: {caminho_planilha}")
    return con
