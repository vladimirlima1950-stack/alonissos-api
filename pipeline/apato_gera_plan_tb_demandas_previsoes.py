import os
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers

def run(pasta_cliente):

    pasta_saida = os.path.join(pasta_cliente, "saida")
    os.makedirs(pasta_saida, exist_ok=True)

    caminho_duckdb = os.path.join(pasta_cliente, "processamento", "previsao.duckdb")

    con = duckdb.connect(caminho_duckdb, read_only=False)

    # ============================================================
    # CONSULTA AO BANCO DE DADOS
    # ============================================================
    df = con.execute("SELECT * FROM tb_dmd_fcst_res_fim").df()

    # ============================================================
    # Cabeçalhos e campos (mantidos exatamente como no PHP)
    # ============================================================
    titulos = [
        'frequência de pedidos', 'importância no faturamento', 'sku número',
        'corr_menos24','corr_menos23','corr_menos22','corr_menos21','corr_menos20','corr_menos19','corr_menos18',
        'corr_menos17','corr_menos16','corr_menos15','corr_menos14','corr_menos13','corr_menos12','corr_menos11',
        'corr_menos10','corr_menos9','corr_menos8','corr_menos7','corr_menos6','corr_menos5','corr_menos4','corr_menos3',
        'corr_menos2','corr_menos1','mês corrente','corr_mais1','corr_mais2','corr_mais3','corr_mais4','corr_mais5',
        'corr_mais6','corr_mais7','corr_mais8','corr_mais9','corr_mais10','corr_mais11'
    ]

    campos = [
        'class_pedidos','class_valores','sku',
        'corr_menos24','corr_menos23','corr_menos22','corr_menos21','corr_menos20','corr_menos19','corr_menos18',
        'corr_menos17','corr_menos16','corr_menos15','corr_menos14','corr_menos13','corr_menos12','corr_menos11',
        'corr_menos10','corr_menos9','corr_menos8','corr_menos7','corr_menos6','corr_menos5','corr_menos4','corr_menos3',
        'corr_menos2','corr_menos1','corr','corr_mais1','corr_mais2','corr_mais3','corr_mais4','corr_mais5',
        'corr_mais6','corr_mais7','corr_mais8','corr_mais9','corr_mais10','corr_mais11'
    ]

    # ============================================================
    # CRIAÇÃO DA PLANILHA
    # ============================================================
    wb = Workbook()
    ws = wb.active

    # Cabeçalhos
    for idx, titulo in enumerate(titulos, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = titulo
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 15

    # ============================================================
    # LINHAS DE DADOS
    # ============================================================
    linha = 2

    for _, row in df.iterrows():
        for col_idx, campo in enumerate(campos, start=1):

            valor = row[campo]
            cell = ws.cell(row=linha, column=col_idx)

            # Alinhamento igual ao PHP
            if campo == "sku":
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

            # Valor
            if isinstance(valor, (int, float)):
                cell.value = float(valor)
            else:
                cell.value = valor

            # Formatação numérica igual ao PHP
            if "corr" in campo or isinstance(valor, (int, float)):
                cell.number_format = numbers.FORMAT_NUMBER

        linha += 1

    # ============================================================
    # SALVA ARQUIVO
    # ============================================================
    caminho_arquivo = os.path.join(pasta_saida, "tabela_demandas_previsoes.xlsx")
    wb.save(caminho_arquivo)

    print(f"Planilha gerada com sucesso em: {caminho_arquivo}")

    return caminho_arquivo
