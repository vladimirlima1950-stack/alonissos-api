import os
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers

def run(pasta_cliente):

    pasta_saida = os.path.join(pasta_cliente, "saida")
    os.makedirs(pasta_saida, exist_ok=True)

    caminho_duckdb = os.path.join(pasta_cliente, "processamento", "previsao.duckdb")

    con = duckdb.connect(caminho_duckdb, read_only=False)

    # ============================================================
    # CONSULTA AO BANCO DE DADOS
    # ============================================================
    df = con.execute("""
        SELECT classe_de_pedidos, classe_de_custo, qtde_skus, porc_skus,
               qtde_pedidos, porc_pedidos, custo_12meses, porc_custo_12meses,
               custo_estoque, porc_custo_estoque, giro_do_estoque
        FROM tb_apres2_fim
    """).df()

    # ============================================================
    # CABEÇALHOS E CAMPOS (mantidos exatamente como no PHP)
    # ============================================================
    colunas = [
        'frequência de pedidos',
        'importância no faturamento',
        'quantidade de diferentes produtos',
        'porcentagem do produto',
        'quantidade de pedidos em 12 meses',
        'porcentagem de pedidos em 12 meses',
        'custo do produto em 12 meses',
        'porcentagem do custo em 12 meses',
        'custo do estoque',
        'porcentagem do custo do estoque',
        'giro do estoque'
    ]

    campos = [
        'classe_de_pedidos',
        'classe_de_custo',
        'qtde_skus',
        'porc_skus',
        'qtde_pedidos',
        'porc_pedidos',
        'custo_12meses',
        'porc_custo_12meses',
        'custo_estoque',
        'porc_custo_estoque',
        'giro_do_estoque'
    ]

    # ============================================================
    # CRIAÇÃO DA PLANILHA
    # ============================================================
    wb = Workbook()
    ws = wb.active

    # Cabeçalhos
    for idx, titulo in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = titulo
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 15

    # ============================================================
    # PREENCHENDO DADOS
    # ============================================================
    linha = 2
    for _, row in df.iterrows():
        for col_idx, campo in enumerate(campos, start=1):
            valor = row[campo]
            cell = ws.cell(row=linha, column=col_idx)

            # mantém a lógica do PHP
            if isinstance(valor, (int, float)):
                cell.value = float(valor)
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = valor

            # formatação numérica igual ao PHP
            if campo in ('qtde_skus', 'qtde_pedidos'):
                cell.number_format = '0'
            elif 'porc' in campo:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
            elif isinstance(valor, (int, float)):
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        linha += 1

    # ============================================================
    # SALVA ARQUIVO
    # ============================================================
    caminho_arquivo = os.path.join(pasta_saida, "tabela_apres2.xlsx")
    wb.save(caminho_arquivo)

    print(f"Planilha gerada com sucesso em: {caminho_arquivo}")

    return caminho_arquivo
